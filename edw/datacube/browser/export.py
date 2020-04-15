import cairosvg
import csv
import datetime
import json
import logging
import tempfile

from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl import Workbook
from Products.Five.browser import BrowserView
from StringIO import StringIO
from zope.component import queryMultiAdapter

logger = logging.getLogger(__name__)


class ExportCSV(BrowserView):
    """ Export to CSV
    """
    def write_headers(self, sheet, headers, row):
        col = 1
        for h in headers:
            sheet.cell(row=row, column=col).value = h.upper()
            sheet.cell(row=row, column=col).font = Font(bold=True)
            col += 1

    def write_encoded_val(self, headers, sheet, encoded, row):
        for idx, h in enumerate(headers):
            sheet.cell(row=row, column=idx + 1).value = encoded[h]

    def datapoints(self, sheet, chart_data):
        """ Export single dimension series to CSV
        """
        # check if multiple series
        headers = ['category', 'code', 'value', 'note', 'flag']
        if chart_data[0].get('name'):
            headers = ['series'] + headers

        self.write_headers(sheet, headers, row=1)
        sheet.column_dimensions['A'].width = 50

        row = 2
        for series in chart_data:
            for point in series['data']:
                value = (
                    point.get('y')
                    if not point.get('isNA', False)
                    else None
                )
                flag = point.get('attributes').get('flag')
                encoded = {
                    'series': series.get('name'),
                    'category': point.get('name'),
                    'code': point.get('code'),
                    'value': value,
                    'flag': "{} ({})".format(
                        flag.get('notation'), flag.get('label')
                    ) if flag else None,
                    'note': point.get('attributes').get('note'),
                }
                self.write_encoded_val(headers, sheet, encoded, row)
                row += 1

    def datapoints_n(self, sheet, chart_data):
        """ Export multiple dimension series to CSV
        """
        keys = set(chart_data[0][0].get('data', [{}])[0].keys())

        headers = ['series', 'name', 'x', 'y']
        if 'z' in keys:
            headers.append('z')
        headers += ['note', 'flag']
        self.write_headers(sheet, headers, row=1)

        row = 2
        for series in chart_data:
            for point in series:
                point_data = point['data'][0]
                attributes = point_data.get('attributes', {})
                flag = attributes.get('flag', {}).get('x', {})
                flag = "{} ({})".format(
                    flag.get('notation'), flag.get('label')
                ) if flag else None
                encoded = {
                    'series': point['name'],
                    'name': point_data['name'],
                    'x': point_data['x'],
                    'y': point_data['y'],
                    'z': point_data['z'] if 'z' in keys else None,
                    'flag': flag,
                    'note': attributes.get('note', {}).get('x', None),
                }

                self.write_encoded_val(headers, sheet, encoded, row)
                row += 1

    def datapoints_profile(self, sheet, chart_data):
        headers = [
            'period', 'indicator', 'EU average', 'value', 'note', 'flag'
        ]
        self.write_headers(sheet, headers, row=1)

        row = 2
        for series in chart_data:
            for point in series['data']:
                attr = point.get('attributes', {})
                flag = attr.get('flag')
                encoded = {
                    'period': attr['time-period']['notation'],
                    'indicator': point.get('name'),
                    'EU average':  point.get('eu'),
                    'value': point.get('original'),
                    'flag': '{} ({})'.format(
                            flag.get('notation'), flag.get('label')
                    ) if flag else None,
                    'note': attr.get('note')
                }
                self.write_encoded_val(headers, sheet, encoded, row)
                row += 1

    def datapoints_profile_table(self, sheet, chart_data):
        row = 1
        for series in chart_data:
            encoded = {}
            latest = series['data']['latest']
            years = [
                '%s' % (latest-3),
                '%s' % (latest-2),
                '%s' % (latest-1),
                '%s' % (latest)
            ]

            headers = (['country', 'indicator', 'breakdown', 'unit'] + years +
                       ['EU28 value %s' % latest, 'rank'])

            self.write_headers(sheet, headers, row=row)
            sheet.column_dimensions['A'].width = 15
            sheet.column_dimensions['E'].width = 12
            sheet.column_dimensions['F'].width = 12
            sheet.column_dimensions['G'].width = 12
            sheet.column_dimensions['H'].width = 12

            row += 1
            encoded['country'] = series['data']['ref-area']['label']
            for ind in series['data']['table'].values():
                encoded['indicator'] = ind['indicator']
                encoded['breakdown'] = ind['breakdown']
                encoded['unit'] = ind['unit-measure']
                for year in years:
                    encoded[year] = ind.get(year, '-')
                encoded['EU28 value %s' % latest] = ind.get('eu', '-')
                rank = ind.get('rank', '-')
                encoded['rank'] = rank if rank != 0 else '-'

                self.write_encoded_val(headers, sheet, encoded, row)
                row += 1

    def datapoints_profile_polar(self, sheet, chart_data):
        headers = ['country', 'category', 'indicator', 'breakdown',
                   'unit', 'eu', 'original', 'period']

        self.write_headers(sheet, headers, row=1)

        row = 2
        for series in chart_data:
            for point in series['data']:
                attrs = point['attributes']
                encoded = {
                 'country': attrs['ref-area']['notation'],
                 'category': point['title'],
                 'indicator': attrs['indicator']['notation'],
                 'breakdown': attrs['breakdown']['notation'],
                 'unit': attrs['unit-measure']['notation'],
                 'eu': attrs['eu'],
                 'original': attrs['original'],
                 'period': attrs['time-period']['notation'],
                }

                self.write_encoded_val(headers, sheet, encoded, row)
                row += 1

    def write_general_sheet(self, sheet, data):
        sheet.column_dimensions['B'].width = 100
        sheet.cell(row=1, column=1).value = 'Chart title'

        chart_url = data.get('chart-url', None)
        chart_title = data.get('chart_title', '-')
        chart_subtitle = data.get('chart_subtitle', None)

        if chart_subtitle:
            chart_title = '{} | {}'.format(chart_title, chart_subtitle)

        if chart_url:
            sheet.cell(row=1, column=2).hyperlink = chart_url
            sheet.cell(row=1, column=2).style = "Hyperlink"
        sheet.cell(row=1, column=2).value = chart_title

        sheet.cell(row=2, column=1).value = 'Source dataset'
        sheet.cell(row=2, column=2).value = data.get('source-dataset', '-')

        sheet.cell(row=3, column=1).value = 'Dataset metadata and download'
        sheet.cell(row=3, column=2).hyperlink = data.get('metadata_details_url')
        sheet.cell(row=3, column=2).value = data.get('metadata_details_url')
        sheet.cell(row=3, column=2).style = "Hyperlink"

        sheet.cell(row=4, column=1).value = 'List of available indicators'
        sheet.cell(row=4, column=2).hyperlink = data.get('indicators_details_url')
        sheet.cell(row=4, column=2).value = data.get('indicators_details_url')
        sheet.cell(row=4, column=2).style = "Hyperlink"

        sheet.cell(row=5, column=1).value = 'Extraction date'
        sheet.cell(row=5, column=2).value = datetime.date.today()

    def write_filter_row(self, sheet, row, dimension, notation):
        cube = self.context.get_cube()
        notation_meta = cube.metadata.lookup_notation(dimension, notation)
        sheet.cell(row=row, column=2).value = notation
        sheet.cell(row=row, column=3).value = notation_meta.get('label', notation)
        sheet.cell(row=row, column=4).value = notation_meta.get('definition')
        sheet.cell(row=row, column=5).value = notation_meta.get('source_label')

    def write_applied_filters_sheet(self, sheet, data):
        header = ['Filter', 'Notation', 'Label', 'Definition', 'Source']
        self.write_headers(sheet, header, row=1)
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 60
        sheet.column_dimensions['D'].width = 60
        sheet.column_dimensions['E'].width = 60

        cube = self.context.get_cube()
        row = 2
        for dimension, value in data:
            # get dimension names
            if (
                dimension.endswith('-slider-values') or
                dimension.endswith('-normalized-values')
            ):
                # DESI sliders need special handling
                # TODO
                continue
            dimension_prefix = None
            if dimension[:2] in ('x-', 'y-', 'z-'):
                # multi-dimensional chart
                dimension_prefix = dimension[0].upper()
                dimension = dimension[2:]
            dimension_uri = cube.metadata.lookup_dimension_uri(dimension)
            dimension_label = (
                cube.metadata.lookup_dimension_info(dimension_uri)
                    .get('label', dimension)
            )
            if dimension_prefix:
                dimension_label = '({}) {}'.format(
                    dimension_prefix, dimension_label
                )
            if type(value) is list:
                if (
                    'any' in value or
                    cube.metadata.is_group_dimension(dimension_uri)
                ):
                    # skip redundant selections
                    continue
                first_row = row
                sheet.cell(row=row, column=1).value = dimension_label
                for notation in value:
                    self.write_filter_row(sheet, row, dimension, notation)
                    row += 1
                sheet.merge_cells(
                    start_row=first_row, start_column=1,
                    end_row=row-1, end_column=1
                )
            elif value != 'any':
                # this is a simple value
                sheet.cell(row=row, column=1).value = dimension_label
                self.write_filter_row(sheet, row, dimension, value)
                row += 1

    def write_observation_data_sheet(self, sheet, chart_data):
        formatters = {
            'scatter': self.datapoints_n,
            'bubbles': self.datapoints_n,
            'country_profile_bar': self.datapoints_profile,
            'country_profile_table': self.datapoints_profile_table,
            'country_profile_polar_polar': self.datapoints_profile_polar,
        }

        chart_type = self.request.form.pop('chart_type')

        formatter = formatters.get(chart_type, self.datapoints)
        formatter(sheet, chart_data)

    def make_sheet(self, wb, name, data, sheet_fct):
        last_sheet = name == 'Data'
        filters_sheet = name == 'Applied Filters'

        if not (last_sheet or filters_sheet):
            sheet = wb.active
            sheet.title = name
        else:
            sheet = wb.create_sheet(name)

        sheet_fct(sheet, data)
        self.apply_styles(sheet)

    @staticmethod
    def apply_styles(sheet, min_width=30, alignment='top'):
        # Set cell width, center cells
        for col in sheet.columns:
            idx = col[0].column
            if not sheet.column_dimensions[idx].width:
                sheet.column_dimensions[idx].width = min_width
            for cell in col:
                cell.alignment = Alignment(
                    horizontal='left',
                    vertical=alignment,
                    wrapText=True
                )

    def export(self):
        if not self.request.form.get('chart_data'):
            return

        metadata = json.loads(self.request.form.get('metadata'))
        cube_url = self.request.form.get('cube_url')

        # These are also found in metadata, but may not be always updated
        chart_title = self.request.form.get('chart_title', '')
        chart_subtitle = self.request.form.get('chart_subtitle', '')

        general_info_data = {
            u'chart_title': chart_title,
            u'chart_subtitle': chart_subtitle,
            u'chart-url': metadata['chart-url'],
            u'source-dataset': metadata['source-dataset'],
            u'indicators_details_url': cube_url + '/indicators',
            u'metadata_details_url': cube_url + '/#download',
        }

        wb = Workbook()
        self.make_sheet(
            wb, 'General Information', general_info_data,
            self.write_general_sheet
        )
        self.make_sheet(
            wb, 'Applied Filters', metadata['filters-applied'],
            self.write_applied_filters_sheet
        )
        self.make_sheet(
            wb, 'Data',
            json.loads(self.request.form.pop('chart_data')),
            self.write_observation_data_sheet
        )

        return self.download_xls(wb)

    def convert_to_csv(self, workbook, stream):
        sheets = workbook.worksheets

        writer = csv.writer(stream)
        for sheet in sheets:
            for row in sheet.rows:
                data_row = map(lambda x: x.value, row)
                writer.writerow(data_row)

        return stream

    def download_xls(self, wb):
        to_xlsx = self.request.form.get('format') == 'xlsx'
        title = self.context.getId().replace(" ", "_")
        timestamp = datetime.date.today().strftime("%d_%m_%Y")

        filename = title + '_' + str(timestamp)

        if to_xlsx:
            stream = StringIO()
            wb.save(stream)
            stream.seek(0)

            self.request.response.setHeader(
                'Content-Type',
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            self.request.response.setHeader(
                'Content-Disposition',
                'attachment; filename="%s.xlsx"' % filename)

            return stream.read()
        else:
            output_stream = self.request.response

            self.request.response.setHeader(
                'Content-Type', 'application/csv; charset=utf-8')
            self.request.response.setHeader(
                'Content-Disposition',
                'attachment; filename="%s.csv"' % filename)

            self.convert_to_csv(wb, output_stream)
            return self.request.response


class ExportRDF(BrowserView):
    """ Export to RDF
    """
    def datapoints(self, points):
        """ xxx
        """
        return "Not implemented error"

    def datapoints_xy(self, points):
        """
        """
        return "Not implemented error"

    def export(self):
        """ Export to csv
        """
        options = self.request.form.get('options', "{}")
        method = self.request.form.get('method', 'datapoints')
        formatter = getattr(self, method, None)

        self.request.form = json.loads(options)
        points = queryMultiAdapter((self.context, self.request), name=method)

        self.request.response.setHeader(
            'Content-Type', 'application/rdf+xml')
        self.request.response.setHeader(
            'Content-Disposition',
            'attachment; filename="%s.rdf"' % self.context.getId())

        if not points:
            return ""

        if formatter:
            return formatter(points)

        return ""


def cairosvg_surface_color(string=None, opacity=1):
    """
    Replace ``string`` representing a color by a RGBA tuple.
    Function overwritten to catch exceptions at line 295.
    """
    from cairosvg.surface.colors import COLORS

    if not string or string in ("none", "transparent"):
        return (0, 0, 0, 0)

    string = string.strip().lower()

    if string in COLORS:
        string = COLORS[string]

    if string.startswith("rgba"):
        r, g, b, a = tuple(
            float(i.strip(" %")) * 2.55 if "%" in i else float(i)
            for i in string.strip(" rgba()").split(","))
        return r / 255, g / 255, b / 255, a * opacity
    elif string.startswith("rgb"):
        r, g, b = tuple(
            float(i.strip(" %")) / 100 if "%" in i else float(i) / 255
            for i in string.strip(" rgb()").split(","))
        return r, g, b, opacity

    if len(string) in (4, 5):
        string = "#" + "".join(2 * char for char in string[1:])
    if len(string) == 9:
        try:
            opacity *= int(string[7:9], 16) / 255
        except:
            pass

    try:
        plain_color = tuple(
            int(value, 16) / 255. for value in (
                string[1:3], string[3:5], string[5:7]))
    except ValueError:
        # Unknown color, return black
        return (0, 0, 0, 1)
    else:
        return plain_color + (opacity,)

class ExportSvg(BrowserView):
    def export(self):
        """
        Simply serve the png submitted as request parameter.
        This is needed only for the Content-Disposition header.
        """
        svg = self.request.get('svg')
        filename = self.request.get('filename', 'chart');
        self.request.response.setHeader(
            'Content-Type', 'image/svg+xml')
        self.request.response.setHeader(
            'Content-Disposition',
            'attachment; filename="' + filename + '.svg"')
        self.request.response.write(svg)
        return self.request.response

class SvgToPng(BrowserView):
    def convert(self):
        """
        Converts a svg to png and http returns the png.
        """
        svg = self.request.get('svg')
        filename = self.request.get('filename', 'chart');
        png_file = tempfile.TemporaryFile(mode='w+b')

        cairosvg.surface.color = cairosvg_surface_color
        cairosvg.svg2png(bytestring=svg, write_to=png_file)

        self.request.response.setHeader(
            'Content-Type', 'image/png')
        self.request.response.setHeader(
            'Content-Disposition',
            'attachment; filename="' + filename + '.png"')

        png_file.flush()
        png_file.seek(0)

        self.request.response.write(png_file.read())

        return self.request.response


class UnicodeWriter(object):
    def __init__(self, *args, **kwargs):
        self.writer = csv.writer(*args, **kwargs)

    def writerow(self, row):
        self.writer.writerow([
            x.encode('utf-8') if isinstance(x, unicode) else x for x in row])
